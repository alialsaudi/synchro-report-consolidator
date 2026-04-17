"""Write a consolidated Synchro-report .xlsx from parsed intersections.

Mirrors the VBA macro's 8-rows-per-intersection block layout (see
src/vba/Main.bas around line 240) so the output is recognizable to existing
users. Each input folder becomes one sheet; intersections stack vertically.

Runtime dependency: openpyxl.

Usage
-----
Library:
    from synchro_parser import parse_file
    from synchro_writer import write_consolidated, parse_folder

    datasets = [parse_folder("path/to/2019 MD")]
    write_consolidated("out.xlsx", datasets)

CLI:
    python synchro_writer.py --out report.xlsx path/to/2019\\ MD path/to/2019\\ PM
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from synchro_parser import Intersection, as_number, parse_file

# (display label, subsection to look in) — order is the row order in each block.
DEFAULT_MOES: list[tuple[str, str]] = [
    ("Level of Service", "HCM Signalized Intersection Capacity Analysis"),
    ("Delay (s)", "HCM Signalized Intersection Capacity Analysis"),
    ("v/c Ratio", "HCM Signalized Intersection Capacity Analysis"),
    ("Queue Length 95th (m)", "Queues"),
    ("Storage Length (m)", "Lanes and Geometrics"),
]

BLOCK_ROWS = 1 + 1 + len(DEFAULT_MOES) + 1  # title + direction header + MOE rows + blank separator
TITLE_FILL = PatternFill("solid", fgColor="4472C4")  # Excel Accent1-ish
TITLE_FONT = Font(bold=True, color="FFFFFF")
HEADER_FONT = Font(bold=True)
CENTERED = Alignment(horizontal="center", vertical="center")


def parse_folder(folder: str | Path) -> tuple[str, list[Intersection]]:
    """Parse every .txt in `folder` and return (sheet_name, intersections).

    Sheet name is the folder's basename, uppercased to match the VBA convention.
    Intersections from multiple files are concatenated in filename order. No
    deduplication — same intersection number in two files produces two blocks.
    """
    p = Path(folder)
    if not p.is_dir():
        raise NotADirectoryError(p)
    all_isx: list[Intersection] = []
    for txt in sorted(p.glob("*.txt")):
        all_isx.extend(parse_file(txt))
    return p.name.upper(), all_isx


def write_consolidated(
    output_path: str | Path,
    datasets: list[tuple[str, list[Intersection]]],
    moes: list[tuple[str, str]] | None = None,
) -> None:
    """Write one sheet per dataset into a new .xlsx at `output_path`."""
    moes = moes or DEFAULT_MOES
    wb = Workbook()
    # openpyxl creates a default sheet; we'll reuse it for the first dataset.
    first = True
    for sheet_name, intersections in datasets:
        safe_name = _safe_sheet_name(sheet_name)
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = safe_name
        ws.column_dimensions["A"].width = 15.71
        ws.column_dimensions["B"].width = 21
        _write_sheet(ws, intersections, moes)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_sheet(
    ws,
    intersections: list[Intersection],
    moes: list[tuple[str, str]],
) -> None:
    for idx, isx in enumerate(intersections):
        base = 1 + idx * BLOCK_ROWS  # openpyxl is 1-indexed

        # Row 0 of block: "Intersection N: Name" — spans direction columns
        directions = _directions_for_intersection(isx, moes)
        total_cols = 2 + max(1, len(directions))  # A (label) + B (row title) + direction cols
        title = f"Intersection {isx.number}: {isx.name}"
        ws.cell(row=base, column=1, value=title)
        if total_cols > 1:
            ws.merge_cells(
                start_row=base, start_column=1, end_row=base, end_column=total_cols
            )
        title_cell = ws.cell(row=base, column=1)
        title_cell.fill = TITLE_FILL
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTERED

        # Row 1: "" | "Direction" | EBL | EBT | ...
        ws.cell(row=base + 1, column=2, value="Direction").font = HEADER_FONT
        for j, d in enumerate(directions):
            c = ws.cell(row=base + 1, column=3 + j, value=d)
            c.font = HEADER_FONT
            c.alignment = CENTERED

        # Row 2: "Data Set" label in A
        ws.cell(row=base + 2, column=1, value="Data Set").font = HEADER_FONT

        # MOE rows: B column is the row title; C+ are values per direction.
        for k, (label, subsection) in enumerate(moes):
            r = base + 2 + k
            ws.cell(row=r, column=2, value=label).font = HEADER_FONT
            for j, d in enumerate(directions):
                raw = isx.get_metric(label, d, subsection=subsection)
                if raw is None:
                    continue
                value = as_number(raw)
                c = ws.cell(row=r, column=3 + j, value=value)
                c.alignment = CENTERED


def _directions_for_intersection(
    isx: Intersection, moes: list[tuple[str, str]]
) -> list[str]:
    """Union of direction codes across the subsections we'll read, preserving
    declaration order from each subsection. This matches the VBA behavior of
    taking the direction columns from `Lane Group`/`Movement` header rows,
    while tolerating subsections that omit some directions (e.g. Queues has
    no U-turns in Synchro 11 samples)."""
    seen: list[str] = []
    for _, subsection_name in moes:
        sub = isx.subsections.get(subsection_name)
        if sub is None:
            continue
        for d in sub.directions:
            if d not in seen:
                seen.append(d)
    return seen


_INVALID_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if c in _INVALID_SHEET_CHARS else c for c in name)
    return cleaned[:31]  # Excel limit


def _cli(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description="Consolidate Synchro exports into one .xlsx.")
    p.add_argument("folders", nargs="+", help="Input folder(s), each becomes one sheet.")
    p.add_argument("--out", "-o", required=True, help="Output .xlsx path.")
    args = p.parse_args(argv[1:])

    datasets = [parse_folder(f) for f in args.folders]
    total = sum(len(isx_list) for _, isx_list in datasets)
    if total == 0:
        print("no intersections parsed — check that folders contain .txt files", file=sys.stderr)
        return 1
    write_consolidated(args.out, datasets)
    print(f"wrote {args.out}: {len(datasets)} sheet(s), {total} intersection(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(_cli(sys.argv))
